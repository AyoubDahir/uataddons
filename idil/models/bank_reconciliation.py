import base64
import io
import logging
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

class IdilBankStatement(models.Model):
    _name = 'idil.bank.statement'
    _description = 'Bank Statement'
    # _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'date desc, id desc'

    name = fields.Char(string='Reference', required=True, copy=False, readonly=True, default='New')
    date = fields.Date(string='Date', required=True, default=fields.Date.context_today)
    company_id = fields.Many2one('res.company', string='Company', required=True, default=lambda self: self.env.company)
    
    account_id = fields.Many2one(
        'idil.chart.account', 
        string='Bank Account', 
        required=True, 
        domain=[('account_type', 'in', ['bank_transfer', 'cash'])]
    )
    
    balance_start = fields.Float(string='Starting Balance', digits=(16, 2))
    balance_end_real = fields.Float(string='Ending Balance', digits=(16, 2))
    balance_end = fields.Float(string='Computed Balance', compute='_compute_balance_end', store=True)
    difference = fields.Float(compute='_compute_difference', store=True)
    
    line_ids = fields.One2many('idil.bank.statement.line', 'statement_id', string='Statement Lines')
    
    # Excel Import Fields
    excel_file = fields.Binary(string='Upload Excel')
    excel_filename = fields.Char(string='Excel Filename')
    
    template_file = fields.Binary(string='Template File')
    template_filename = fields.Char(string='Template Filename')
    
    state = fields.Selection([
        ('draft', 'Draft'),
        ('open', 'Processing'),
        ('posted', 'Validated'),
        ('cancel', 'Cancelled')
    ], string='Status', default='draft', tracking=True)

    # Discrpancy / Unmatched System View
    unmatched_system_line_ids = fields.One2many('idil.transaction_bookingline', compute='_compute_unmatched_system_lines', string="Unmatched System Transactions")

    @api.model
    def create(self, vals):
        if vals.get('name', 'New') == 'New':
            vals['name'] = self.env['ir.sequence'].next_by_code('idil.bank.statement') or 'New'
        return super(IdilBankStatement, self).create(vals)

    @api.depends('line_ids.amount', 'balance_start')
    def _compute_balance_end(self):
        for stmt in self:
            stmt.balance_end = stmt.balance_start + sum(stmt.line_ids.mapped('amount'))

    @api.depends('balance_end_real', 'balance_end')
    def _compute_difference(self):
        for stmt in self:
            stmt.difference = stmt.balance_end_real - stmt.balance_end

    def _compute_unmatched_system_lines(self):
        BookingLine = self.env['idil.transaction_bookingline']
        for stmt in self:
            if not stmt.account_id:
                stmt.unmatched_system_line_ids = False
                continue
                
            # Find lines for this account, not reconciled, and within relevant date range (optional)
            # For now, showing ALL unreconciled for this account
            domain = [
                ('account_number', '=', stmt.account_id.id),
                ('is_reconciled', '=', False),
                ('company_id', '=', stmt.company_id.id)
            ]
            stmt.unmatched_system_line_ids = BookingLine.search(domain)

    def action_download_template(self):
        """Generate and download a template Excel file"""
        if not openpyxl:
            raise UserError(_("OpenPyXL is not installed."))

        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Bank Statement Template"
        
        # Headers
        headers = ["Date (YYYY-MM-DD)", "Reference/Label", "Amount (Positive=Deposit, Negative=Withdrawal)", "Notes"]
        sheet.append(headers)
        
        # Sample Data
        sheet.append(["2025-01-01", "INV/2025/001", 1500.00, "Customer Payment"])
        sheet.append(["2025-01-02", "BILL/2025/005", -200.50, "Vendor Payment"])
        
        workbook.save(output)
        output.seek(0)
        
        file_content = base64.b64encode(output.read())
        
        # Create a temporary attachment or returns a download action
        # Simpler: Create a wizard or reuse this model to hold the file temporarily
        # For simplicity, we assign it to a field on a singleton or temporary record, but here we can just return a URL if we had a controller.
        # Since we don't have a controller set up, let's just save it to the current record (if ID exists) or raise info.
        # Best way in Odoo backend without controller: Write to a Transient Model and open a view to download it.
        
        # Let's assume we maintain a 'template_file' field on the main model for simplicity, or we create a transient model.
        # I'll add 'template_file' and 'template_filename' to the model.
        
        self.write({
            'template_file': file_content,
            'template_filename': 'Bank_Statement_Template.xlsx'
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/?model=idil.bank.statement&id=%s&field=template_file&download=true&filename=Bank_Statement_Template.xlsx' % self.id,
            'target': 'self',
        }

    def action_import_excel(self):
        """Import lines from Excel and Auto-Match"""
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_("Please upload an Excel file first."))

        try:
            file_data = base64.b64decode(self.excel_file)
            workbook = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            sheet = workbook.active
            
            lines_to_create = []
            row_idx = 0
            for row in sheet.iter_rows(values_only=True):
                row_idx += 1
                if row_idx == 1: continue # Skip header
                
                # Check for empty row
                if not row or not row[0]: continue
                
                date_val = row[0]
                ref_val = str(row[1]) if row[1] else 'Exel Import'
                amount_val = row[2]
                note_val = str(row[3]) if len(row) > 3 and row[3] else ''

                if not isinstance(amount_val, (int, float)):
                    try:
                        amount_val = float(amount_val)
                    except:
                        continue 
                
                if hasattr(date_val, 'date'):
                    date_val = date_val.date()
                
                lines_to_create.append({
                    'statement_id': self.id,
                    'date': date_val,
                    'payment_ref': ref_val,
                    'amount': amount_val,
                    'note': note_val,
                    'match_status': 'unmatched'
                })
            
            if lines_to_create:
                # Remove existing lines if re-importing? 
                # For now append. User can clear manually if needed.
                self.env['idil.bank.statement.line'].create(lines_to_create)
            
            # Auto Run Matching
            self.action_auto_match_advanced()
                
            self.excel_file = False 
            
        except Exception as e:
             raise UserError(_("Error importing Excel: %s") % str(e))

    def action_open(self):
        self.write({'state': 'open'})

    def action_validate(self):
        for stmt in self:
            if abs(stmt.difference) > 0.001:
                raise UserError(_('The ending balance is incorrect. Please check the difference.'))
            if any(line.match_status != 'matched' for line in stmt.line_ids):
                 # We allow validate if user insists? usually strict.
                 # User asked to highlight those not matched.
                 # Let's stricter:
                 unmatched = stmt.line_ids.filtered(lambda l: l.match_status != 'matched')
                 if unmatched:
                     raise UserError(_('There are %s unmatched or mismatching lines. Please resolve them.') % len(unmatched))
                     
        self.write({'state': 'posted'})

    def action_cancel(self):
        self.write({'state': 'cancel'})

    def action_auto_match_advanced(self):
        """
        Advanced Matching:
        1. Exact Match: Ref + Amount + Date (Relaxed date?) -> Matched
        2. Ref Match + Date Match BUT Amount Diff -> Mismatch (Highlighted)
        """
        self.ensure_one()
        BookingLine = self.env['idil.transaction_bookingline']
        Match = self.env['idil.reconciliation.match']
        
        # We assume 'payment_ref' in statement maps to 'transaction_number' or 'ref' in booking
        
        for st_line in self.line_ids.filtered(lambda l: not l.is_reconciled):
            
            # Base Domain: same account, unreconciled
            base_domain = [
                ('account_number', '=', self.account_id.id),
                ('is_reconciled', '=', False),
                ('company_id', '=', self.company_id.id)
            ]
            
            # Helper to find by Ref
            # We search for booking lines where the transaction header's number matches the excel ref
            # This is a bit complex in domain, easier to search potential candidates by amount/date first?
            # Or search by Ref first? Ref is strongest key.
            
            # Let's try matching by REF first (if Ref exists)
            candidates = BookingLine.browse()
            if st_line.payment_ref:
                # Search transaction_booking_ids with this ref
                # 'transaction_booking_id.transaction_number'
                # or 'description'
                ref_domain = base_domain + [
                    '|', 
                    ('transaction_booking_id.transaction_number', 'ilike', st_line.payment_ref),
                    ('description', 'ilike', st_line.payment_ref)
                ]
                candidates = BookingLine.search(ref_domain)
            
            match_found = False
            
            # 1. Check Candidates for Exact Match (Amount & Date)
            for cand in candidates:
                # Check Amount
                cand_amount = cand.dr_amount - cand.cr_amount
                # Bank: +Amount = Debit in Ledger? No.
                # If Bank Statement says +100 (Deposit), we recvd money. Ledger should have Debit 100 to Bank.
                # So Bank Amt (+100) == Ledger Dr (100). 
                # Bank Amt (-50) == Ledger Cr (50).
                
                # Normalized Ledger Amount
                ledger_val = cand.dr_amount if cand.dr_amount > 0 else -cand.cr_amount
                
                amount_match = abs(st_line.amount - ledger_val) < 0.01
                date_match = (cand.transaction_date == st_line.date)
                
                if amount_match:
                    # Perfect Match (Ref + Amount). Date is secondary but good to check.
                    # Create Match
                    Match.create({
                        'statement_line_id': st_line.id,
                        'booking_line_id': cand.id,
                        'amount': abs(st_line.amount),
                        'match_type': 'auto'
                    })
                    st_line.match_status = 'matched'
                    match_found = True
                    break
                elif date_match and not amount_match:
                    # Ref + Date match, but Amount mismatch
                    st_line.match_status = 'mismatch'
                    st_line.note = (st_line.note or '') + " [Mismatch: Sys Amount %s]" % ledger_val
                    match_found = True # Handled as mismatch
                    break
            
            if match_found:
                st_line._compute_is_reconciled()
                continue
                
            # 2. Fallback: If no Ref match, try Amount + Date match (Weak match)
            # Only if Ref was empty or yielded no results
            if not match_found:
                # Try finding by Amount + Date
                amt_domain = base_domain + [
                    ('transaction_date', '=', st_line.date)
                ]
                if st_line.amount > 0:
                    amt_domain.append(('dr_amount', '=', st_line.amount))
                else:
                    amt_domain.append(('cr_amount', '=', abs(st_line.amount)))
                
                weak_candidate = BookingLine.search(amt_domain, limit=1)
                
                if weak_candidate:
                     Match.create({
                        'statement_line_id': st_line.id,
                        'booking_line_id': weak_candidate.id,
                        'amount': abs(st_line.amount),
                        'match_type': 'auto'
                    })
                     st_line.match_status = 'matched'
                     st_line._compute_is_reconciled()


class IdilBankStatementLine(models.Model):
    _name = 'idil.bank.statement.line'
    _description = 'Bank Statement Line'

    statement_id = fields.Many2one('idil.bank.statement', string='Statement', required=True, ondelete='cascade')
    date = fields.Date(string='Date', required=True)
    payment_ref = fields.Char(string='Ref/Label', required=True)
    partner_id = fields.Many2one('idil.vendor.registration', string='Partner') 
    amount = fields.Float(string='Amount', digits=(16, 2))
    note = fields.Text(string='Notes')
    
    match_ids = fields.One2many('idil.reconciliation.match', 'statement_line_id', string='Matches')
    is_reconciled = fields.Boolean(compute='_compute_is_reconciled', store=True)
    
    match_status = fields.Selection([
        ('unmatched', 'Unmatched'),
        ('matched', 'Matched'),
        ('mismatch', 'Mismatch (Amount)')
    ], string='Match Status', default='unmatched', store=True)

    @api.depends('match_ids', 'amount')
    def _compute_is_reconciled(self):
        for line in self:
            matched_amt = sum(line.match_ids.mapped('amount'))
            is_rec = abs(abs(line.amount) - matched_amt) < 0.001
            line.is_reconciled = is_rec
            if is_rec:
                line.match_status = 'matched'
            # If not reconciled, status remains whatever it was (unmatched or mismatch)

    def action_open_matches_view(self):
        """Open the form view to manage matches"""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'idil.bank.statement.line',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new', 
        }

    def action_manual_reconciliation(self):
        """Open wizard to manually select a system transaction"""
        self.ensure_one()
        return {
            'name': 'Manual Match',
            'type': 'ir.actions.act_window',
            'res_model': 'idil.bank.reconciliation.manual.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_statement_line_id': self.id}
        }

class IdilReconciliationMatch(models.Model):
    _name = 'idil.reconciliation.match'
    _description = 'Reconciliation Match'

    statement_line_id = fields.Many2one('idil.bank.statement.line', string='Statement Line', required=True, ondelete='cascade')
    booking_line_id = fields.Many2one('idil.transaction_bookingline', string='Journal Item', required=True, ondelete='cascade')
    amount = fields.Float(string='Matched Amount', required=True)
    match_type = fields.Selection([('auto', 'Auto'), ('manual', 'Manual')], default='manual')

    @api.model
    def create(self, vals):
        res = super(IdilReconciliationMatch, self).create(vals)
        res.booking_line_id._compute_reconciliation_status()
        return res

    def unlink(self):
        lines = self.mapped('booking_line_id')
        res = super(IdilReconciliationMatch, self).unlink()
        lines._compute_reconciliation_status()
        return res

class TransactionBookingLineExtension(models.Model):
    _inherit = 'idil.transaction_bookingline'

    statement_line_id = fields.Many2one('idil.bank.statement.line', string='Matched Statement Line', compute='_compute_reconciliation_status', store=True)
    reconciliation_match_ids = fields.One2many('idil.reconciliation.match', 'booking_line_id', string='Reconciliations')
    is_reconciled = fields.Boolean(compute='_compute_reconciliation_status', store=True, string='Reconciled')
    
    @api.depends('reconciliation_match_ids')
    def _compute_reconciliation_status(self):
        for line in self:
            total_matched = sum(line.reconciliation_match_ids.mapped('amount'))
            line_amt = line.dr_amount if line.dr_amount > 0 else line.cr_amount
            line.is_reconciled = abs(line_amt - total_matched) < 0.001
            
            if line.reconciliation_match_ids:
                 # line.statement_line_id = line.reconciliation_match_ids[0].statement_line_id
                 pass
            else:
                 # line.statement_line_id = False
                 pass

class IdilBankManualReconciliationWizard(models.TransientModel):
    _name = 'idil.bank.reconciliation.manual.wizard'
    _description = 'Manual Reconciliation Wizard'

    statement_line_id = fields.Many2one('idil.bank.statement.line', string='Statement Line', required=True, readonly=True)
    account_id = fields.Many2one(related='statement_line_id.statement_id.account_id', readonly=True)
    
    # Filter only unreconciled lines from system
    booking_line_id = fields.Many2one(
        'idil.transaction_bookingline', 
        string='Select System Transaction', 
        required=True,
        domain="[('account_number', '=', account_id), ('is_reconciled', '=', False)]"
    )
    
    amount_to_match = fields.Float(string='Amount to Match', compute='_compute_amount', store=True)

    @api.depends('statement_line_id', 'booking_line_id')
    def _compute_amount(self):
        for wiz in self:
            # Default to statement line remaining amount? Or full amount?
            # Ideally we match full amount for now.
            if wiz.statement_line_id:
                wiz.amount_to_match = abs(wiz.statement_line_id.amount) # Simplification

    def action_match(self):
        self.ensure_one()
        # Create the match
        self.env['idil.reconciliation.match'].create({
            'statement_line_id': self.statement_line_id.id,
            'booking_line_id': self.booking_line_id.id,
            'amount': self.amount_to_match,
            'match_type': 'manual'
        })
        # Update line status
        self.statement_line_id._compute_is_reconciled()
        return {'type': 'ir.actions.act_window_close'}
